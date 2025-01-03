import asyncio
import itertools
from playwright.async_api import async_playwright
import re
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
from PIL import Image as PILImage
import httpx


# Функция для создания таблицы (xlsx файла) из данных от EtsyClient.
# Параметры:
# data - данные от EtsyClient.
# filename - название выходного файла.
# image_size - размер изображения в пикселях, рекомендуется менее 800.

def create_xlsx(data, filename, image_size=(100, 100)):
    # Открытие таблицы
    wb = openpyxl.Workbook()
    ws = wb.active

    # Добавление информации в таблицу
    for word, items in data.items():
        # Добавление ключевого слова
        ws.append([word])
        ws.merge_cells(start_row=ws.max_row, start_column=1,
                       end_row=ws.max_row, end_column=3)
        ws.cell(row=ws.max_row,
                column=1).font = openpyxl.styles.Font(size=14, bold=True)

        # Добавление всех листингов для одного ключевого слова
        for item in items:
            row = ["", item["index"], item["listing_id"]]
            ws.append(row)

            # Изменение размера изображения
            img_data = BytesIO(item["image_data"])
            pil_img = PILImage.open(img_data)
            pil_img = pil_img.resize(image_size)

            img_byte_arr = BytesIO()
            pil_img.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)

            # Добавляем измененное изображение в Excel
            img = ExcelImage(img_byte_arr)
            img.anchor = f"A{ws.max_row}"
            ws.add_image(img)

            # Подгоняем высоту и ширину строк под изображение
            image_height = img.height
            row_height = image_height * 0.75
            ws.row_dimensions[ws.max_row].height = row_height

            column = list(ws.columns)[0][0].column_letter
            ws.column_dimensions[column].width = image_size[0] * 0.075 * 2
        ws.append([])

    # Сохранение таблицы в файл
    wb.save(filename)


def split_list(lst, chunk_size=3):
    return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]


# Основной URL
MAIN_URL = "https://www.etsy.com"


# Класс для работы с Etsy
class EtsyClient():
    # Функция для создание хедера Cookie
    async def get_cookie_header(self):
        cookies = await self._context.cookies()

        cookies_header = "; ".join([f"{cookie['name']}={cookie['value']}"
                                    for cookie in cookies])
        return cookies_header

    # Конструктор класса.
    # Параметры:
    # shop_name - Название магазина.
    # language - Код языка.
    # currency - Код валюты.
    # region - Двухсимвольный код страны.
    # max_page - Глубина сканирования. (Кол-во страниц)
    # file - .txt Файл с ключевыми словами
    def __init__(self, shop_name, language, currency, region, max_page, file, req_cooldown):

        self._max_page = max_page
        self._shop_name = shop_name
        self._language = language
        self._region = region
        self._currency = currency
        self._req_cooldown = req_cooldown

        # Открытие файла с ключевыми словами
        try:
            with open(file, "r") as f:
                self._words = f.readlines()
        except Exception:
            raise Exception(f"Во время загрузки файла {file} произошла ошибка")

    # Функция для создания браузерной сессии на Etsy
    async def init_browser(self):
        # Запуск браузера
        self._playwright = await async_playwright().start()
        self._browser = await self._playwright.chromium.launch(headless=True)

        # Хедеры для всего соединения
        headers = {
            "Accept": ("text/html,application/xhtml+xml,application/xml;"
                       "q=0.9,image/avif,image/webp,image/apng,*/*;"
                       "q=0.8,application/signed-exchange;v=b3;q=0.7"),
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Accept-Language": "ru",
            "Cache-Control": "max-age=0",
            "Downlink": "8.25",
            "Dpr": "1.25",
            "Ect": "4g",
            "Priority": "u=0, i",
            "Referer": "https://www.etsy.com/?dd_referrer=",
            "Origin": "https://www.etsy.com",
            "Rtt": "100",
            "Sec-Ch-Dpr": "1.25",
            "Sec-Ch-Ua": ('"Microsoft Edge";v="131", '
                          '"Chromium";v="131", "Not_A Brand";v="24"'),
            "Sec-Ch-Ua-Arch": '"x86"',
            "Sec-Ch-Ua-Bitness": '"64"',
            "Cookie": "",
            "Sec-Ch-Ua-Full-Version-List": ('"Microsoft Edge";'
                                            'v="131.0.2903.99", '
                                            '"Chromium";'
                                            'v="131.0.6778.140", '
                                            '"Not_A Brand";v="24.0.0.0"'),
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Windows"',
            "Sec-Ch-Ua-Platform-Version": '"15.0.0"',
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "same-origin",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        }

        self._context = await self._browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 "
                "Edg/131.0.0.0"
            ),
            extra_http_headers=headers
        )

        # Открытие страницы для получения куки от капчи
        page = await self._context.new_page()
        await page.goto(MAIN_URL)
        await page.wait_for_load_state('networkidle')
        await page.close()

        # Открытие страницы для получения специальных токенов
        page = await self._context.new_page()
        await page.goto(MAIN_URL)
        await page.wait_for_load_state('networkidle')
        content = await page.content()

        regexp_csrf = (
            r'<meta name=["\']csrf_nonce["\'] '
            r'content=["\'](.+?)["\']'
        )
        match = re.search(regexp_csrf, content)
        if match:
            self._csrf_nonce = match.group(1)

        script_content = await page.locator("script").all_inner_texts()
        page_guid_match = None
        regexp_guid = r'page_guid"\s*[:=]\s*["\'](.+?)["\']'

        for script in script_content:
            page_guid_match = re.search(regexp_guid, script)
            if page_guid_match:
                self._page_guid = page_guid_match.group(1)
                break

        await page.close()

        # Хедеры для api запросов
        self._api_headers = {
            **headers,
            "accept": "*/*",
            "referer": f"{MAIN_URL}/search?q=pyjamas&page=2"
                       "&ref=pagination",
            "x-csrf-token": self._csrf_nonce,
            "x-page-guid": self._page_guid,
            "X-Detected-Locale": "USD|ru|UA",
            "X-Etsy-Protection": "1",
            "X-Requested-With": "XMLHttpRequest",
            "content-type": "application/json",
            "cookie": await self.get_cookie_header(),
        }

        # Смена языка
        api = self._context.request
        await api.post(f"{MAIN_URL}/api/v3/ajax/member/locale-preferences",
                       headers=self._api_headers, data={
                            "currency": self._currency,
                            "language": self._language,
                            "region": self._region
                        })

    # Функция для получения данных с одной страницы по одному ключевому слову
    # Параметры:
    # page - Номер страницы.
    # word - Ключевое слово.
    async def get_data_from_page(self, page, word):
        locale_header = f"{self._currency}|{self._language}|{self._region}"
        self._api_headers = {
                **self._api_headers,
                "X-Detected-Locale": locale_header,
                "cookie": await self.get_cookie_header()}

        payload = {
            "log_performance_metrics": True,
            "specs": {
                "async_search_results": [
                    "Search2_ApiSpecs_WebSearch",
                    {
                        "search_request_params": {
                            "detected_locale": {
                                "language": "ru",
                                "currency_code": self._currency,
                                "region": self._region
                            },
                            "locale": {
                                "language": self._language,
                                "currency_code": self._currency,
                                "region": self._region
                            },
                            "name_map": {
                                "query": "q",
                                "query_type": "qt",
                                "results_per_page": "result_count",
                                "min_price": "min",
                                "max_price": "max"
                            },
                            "parameters": {
                                "q": word,
                                "page": page,
                                "ref": "pagination",
                                "referrer": (f"{MAIN_URL}/uk/search?q={word}"
                                             f"&page={page}&ref=search_bar"),
                                "is_prefetch": True,
                                "placement": "wsg"
                            },
                            "user_id": None
                        },
                        "request_type": "pagination_preact"
                    }
                ]
            },
            "view_data_event_name": "search_single_page_app_specview_rendered",
            "runtime_analysis": False
        }

        # Отправка запроса для получения результатов поиска
        api = self._context.request
        api_path = "/api/v3/ajax/bespoke/member/neu/specs/async_search_results"
        api_url = MAIN_URL + api_path
        response = await api.post(
            api_url,
            data=payload,
            headers=self._api_headers,

        )
        print(response.headers)

        data = await response.json()
        print(data.keys())
        html_data = data["output"]["async_search_results"]
        with open("content.html", "wb") as f:
            f.write(html_data.encode("utf-8"))
        soup = BeautifulSoup(html_data, "html.parser")

        # Получение всех листингов
        listings = soup.select("a.listing-link")
        listings = listings[:-6] if page == 1 else listings

        # Фильтрация листингов по названию магазина
        listings_filtered = [listing for listing in listings
                             if self._shop_name in listing.text]

        # Инициализация клиента для открытия изображений
        client = httpx.AsyncClient()

        # Функция для получения информации об одном листинге
        async def get_listing_data(listing):
            num = listings.index(listing)
            row = num // 4 + 1
            listing_id = listing["data-listing-id"]
            image = listing.select_one("img.wt-image")

            # Получение ссылки на изображение с лучшим качеством
            links = image["srcset"]
            image_link = links.split(", ")[-1].split(" ")[0]
            index = f"{page}.{row:02d}"

            for i in range(10):
                try:
                    image_data = await client.get(image_link)
                    break
                except Exception:
                    continue

            return {
                "image_data": image_data.content,
                "index": index,
                "listing_id": listing_id,
            }

        listings_data = await asyncio.gather(*[get_listing_data(listing)
                                             for listing in listings_filtered])

        return listings_data

    async def get_data_full(self):
        full_data = {}

        for word in self._words:
            word_data = []
            for splited in split_list(list(range(1, self._max_page + 1))):
                print(splited)
                tasks = [self.get_data_from_page(page, word)
                         for page in splited]

                word_data += list(itertools.chain(
                    *await asyncio.gather(*tasks)))

                await asyncio.sleep(self._req_cooldown)
            full_data[word] = word_data

        return full_data

    # Закрытие соединения
    async def close_client(self):
        await self._context.close()
        await self._browser.close()
        await self._playwright.stop()


async def main():
    etsy = EtsyClient("IDlingerieUK", "en-GB", "GBP",
                      "GB", 3, "test.txt", 0)
    await etsy.init_browser()
    data = await etsy.get_data_full()
    await etsy.close_client()
    create_xlsx(data, "test.xlsx")

asyncio.run(main())
